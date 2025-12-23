#!/usr/bin/env python3
"""
Script para procesar archivos Excel de transferencias desde ZIP
Procesa archivos Excel que contienen transferencias realizadas y las sube a Supabase
"""

import sys
import os
import zipfile
import tempfile
import uuid
from datetime import datetime
from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl import load_workbook

# Agregar el directorio raíz al path
sys.path.append(os.path.dirname(__file__))

from src.supabase_client import SupabaseClient
from src.logger import logger

class ExcelTransferProcessor:
    """Procesador de transferencias desde archivos Excel"""
    
    def __init__(self):
        """Inicializa el procesador"""
        self.supabase_client = SupabaseClient()
        logger.info("Procesador de transferencias Excel inicializado")
    
    def extract_zip(self, zip_path: str) -> Optional[str]:
        """
        Extrae el archivo ZIP y retorna la ruta del archivo Excel
        
        Args:
            zip_path: Ruta al archivo ZIP
            
        Returns:
            str: Ruta al archivo Excel extraído, None si hay error
        """
        try:
            logger.info(f"Extrayendo archivo ZIP: {zip_path}")
            
            # Crear directorio temporal
            temp_dir = tempfile.mkdtemp()
            
            # Extraer ZIP
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                logger.info(f"Archivos en el ZIP: {file_list}")
                
                if not file_list:
                    logger.error("El ZIP está vacío")
                    return None
                
                # Extraer todos los archivos
                zip_ref.extractall(temp_dir)
                
                # Buscar archivo Excel
                excel_file = None
                for file in file_list:
                    if file.endswith('.xlsx') or file.endswith('.xls'):
                        excel_file = os.path.join(temp_dir, file)
                        logger.info(f"Archivo Excel encontrado: {excel_file}")
                        break
                
                if not excel_file:
                    logger.error("No se encontró archivo Excel en el ZIP")
                    return None
                
                return excel_file
                
        except Exception as e:
            logger.error(f"Error al extraer ZIP: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def read_excel(self, excel_path: str, start_row: int = 9) -> List[Dict[str, Any]]:
        """
        Lee el archivo Excel y extrae las transferencias
        
        Args:
            excel_path: Ruta al archivo Excel
            start_row: Fila donde empiezan los datos (default: 9, ya que fila 8 es encabezado)
            
        Returns:
            List[Dict]: Lista de transferencias extraídas
        """
        transferencias = []
        
        try:
            logger.info(f"Leyendo archivo Excel: {excel_path}")
            logger.info(f"Los datos empiezan en la fila: {start_row}")
            
            # Cargar workbook
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            logger.info(f"Total de filas en el Excel: {ws.max_row}")
            logger.info(f"Total de columnas en el Excel: {ws.max_column}")
            
            # Leer encabezados de la fila 8 (índice 8 en openpyxl)
            headers = {}
            header_row = ws[8]  # Fila 8 (índice 8, ya que openpyxl es 1-indexed)
            for idx, cell in enumerate(header_row):
                if cell.value:
                    headers[idx] = str(cell.value).strip()
            
            logger.info(f"Encabezados encontrados: {headers}")
            
            # Procesar filas desde start_row (fila 9 en adelante)
            for row_idx in range(start_row, ws.max_row + 1):
                row = ws[row_idx]
                
                # Verificar si la fila está vacía o es una nota
                first_cell = row[0].value
                if not first_cell or (isinstance(first_cell, str) and 'Nota:' in first_cell):
                    logger.info(f"Fila {row_idx} es una nota o está vacía, deteniendo lectura")
                    break
                
                # Extraer datos de la fila
                transferencia = self._extract_row_data(row, headers, row_idx)
                
                if transferencia:
                    transferencias.append(transferencia)
                    logger.info(f"Transferencia extraída de fila {row_idx}: Rastreo={transferencia.get('rastreo')}")
            
            logger.info(f"Total de transferencias extraídas: {len(transferencias)}")
            return transferencias
            
        except Exception as e:
            logger.error(f"Error al leer Excel: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return []
    
    def _extract_row_data(self, row, headers: Dict[int, str], row_num: int) -> Optional[Dict[str, Any]]:
        """
        Extrae los datos de una fila del Excel
        
        Args:
            row: Fila del Excel
            headers: Diccionario con índices de columnas y nombres de encabezados
            row_num: Número de fila (para logging)
            
        Returns:
            Dict: Datos de la transferencia o None si hay error
        """
        try:
            # Mapeo de columnas según lo visto en el Excel
            # Col 1: Fecha de Operación
            # Col 2: Hora de Operación
            # Col 3: Cuenta Origen
            # Col 4: Nombre del Ordenante
            # Col 5: Cuenta Destino
            # Col 7: Banco Destino
            # Col 8: Nombre del Beneficiario
            # Col 12: Transacción
            # Col 13: Importe de la Operación
            # Col 15: Divisa
            # Col 17: Concepto de pago
            # Col 18: Referencia
            # Col 21: No. de Autorización / Clave de Rastreo
            
            # Obtener valores de las celdas
            fecha_operacion = row[1].value if len(row) > 1 else None
            hora_operacion = row[2].value if len(row) > 2 else None
            cuenta_origen = row[3].value if len(row) > 3 else None
            ordenante = row[4].value if len(row) > 4 else None
            cuenta_destino = row[5].value if len(row) > 5 else None
            banco_destino = row[7].value if len(row) > 7 else None
            beneficiario = row[8].value if len(row) > 8 else None
            transaccion = row[12].value if len(row) > 12 else None
            importe = row[13].value if len(row) > 13 else None
            divisa = row[15].value if len(row) > 15 else None
            concepto = row[17].value if len(row) > 17 else None
            referencia = row[18].value if len(row) > 18 else None
            autorizacion_rastreo = row[21].value if len(row) > 21 else None
            
            # Validar que haya datos mínimos
            if not fecha_operacion or not importe:
                logger.warning(f"Fila {row_num} no tiene fecha o importe, omitiendo")
                return None
            
            # Procesar fecha de operación
            fec_operacion = None
            if isinstance(fecha_operacion, datetime):
                fec_operacion = fecha_operacion.strftime('%Y-%m-%d')
            elif isinstance(fecha_operacion, str):
                try:
                    # Intentar parsear diferentes formatos
                    fec_operacion = datetime.strptime(fecha_operacion, '%Y-%m-%d').strftime('%Y-%m-%d')
                except:
                    logger.warning(f"No se pudo parsear la fecha: {fecha_operacion}")
            
            # Procesar hora de operación
            hora_oper = None
            if isinstance(hora_operacion, datetime):
                hora_oper = hora_operacion.strftime('%H:%M:%S')
            elif isinstance(hora_operacion, str):
                hora_oper = hora_operacion.strip()
            
            # Procesar importe
            importe_val = None
            if isinstance(importe, (int, float)):
                importe_val = float(importe)
            elif isinstance(importe, str):
                try:
                    # Limpiar formato de moneda si existe
                    importe_clean = importe.replace('$', '').replace(',', '').strip()
                    importe_val = float(importe_clean)
                except:
                    logger.warning(f"No se pudo parsear el importe: {importe}")
            
            # Procesar moneda
            moneda_val = 'MXN'  # Default
            if divisa:
                if isinstance(divisa, str):
                    divisa_upper = divisa.upper().strip()
                    if divisa_upper == 'MN':
                        moneda_val = 'MXN'
                    elif len(divisa_upper) == 3:
                        moneda_val = divisa_upper
                else:
                    moneda_val = str(divisa).upper().strip()
            
            # Procesar clave de rastreo (puede venir como "2500189013213/BB1863873013213")
            rastreo = None
            autorizacion = None
            if autorizacion_rastreo:
                if isinstance(autorizacion_rastreo, str):
                    # Separar si viene con "/"
                    if '/' in autorizacion_rastreo:
                        parts = autorizacion_rastreo.split('/')
                        autorizacion = parts[0].strip() if len(parts) > 0 else None
                        rastreo = parts[1].strip() if len(parts) > 1 else None
                        # Si la primera parte parece ser un número largo, es autorización
                        # Si la segunda parte empieza con letras (BB, BNET, etc.), es rastreo
                        if rastreo and not rastreo[0].isalpha():
                            # Intercambiar si el formato está al revés
                            temp = rastreo
                            rastreo = autorizacion
                            autorizacion = temp
                    else:
                        # Si no tiene "/", verificar si es rastreo o autorización
                        # Los rastreos suelen empezar con letras (BB, BNET) o tener formato específico
                        if autorizacion_rastreo.strip()[0].isalpha() or len(autorizacion_rastreo.strip()) >= 13:
                            rastreo = autorizacion_rastreo.strip()
                        else:
                            autorizacion = autorizacion_rastreo.strip()
                else:
                    rastreo = str(autorizacion_rastreo).strip()
            
            # Si no hay rastreo pero hay referencia, intentar usar referencia como rastreo temporal
            # (aunque idealmente debería tener rastreo)
            if not rastreo and referencia:
                logger.warning(f"Fila {row_num} no tiene clave de rastreo, se usará solo referencia para validación")
            
            # Crear estructura de datos compatible con movbancarios
            # NOTA: idUnico se genera automáticamente en la BD, no lo enviamos
            transferencia_data = {
                'idmov': str(uuid.uuid4()),
                'fc': datetime.now().isoformat(),
                'asunto': f'Transferencia desde Excel - {fec_operacion or "N/A"}',
                'fecOperacion': fec_operacion,
                'horaOperacion': hora_oper,
                'ordenante': str(ordenante).strip() if ordenante else None,
                'ctaDestino': str(cuenta_destino).strip() if cuenta_destino else None,
                'bcoDestino': str(banco_destino).strip() if banco_destino else None,
                'beneficiario': str(beneficiario).strip() if beneficiario else None,
                'importe': importe_val,
                'moneda': moneda_val,
                'cancepto': str(concepto).strip() if concepto else None,
                'referencia': str(referencia).strip() if referencia else None,
                'rastreo': rastreo,
                'autorizacion': autorizacion,
                'bancoEmisor': None,  # No disponible en el Excel
                'tipo': 'Transferencia SPEI',
                'aplicado': False,
                'manual': True,  # Marcado como manual porque viene de Excel
                'Operacion': str(transaccion).strip() if transaccion else 'Transferencia Interbancaria SPEI'
            }

            # Guardar idUnico solo para validación de duplicados (no se inserta en BD)
            transferencia_data['_idUnico'] = rastreo or str(referencia).strip() if referencia else str(uuid.uuid4())
            
            return transferencia_data
            
        except Exception as e:
            logger.error(f"Error al extraer datos de fila {row_num}: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def process_zip_file(self, zip_path: str) -> Dict[str, Any]:
        """
        Procesa un archivo ZIP con transferencias Excel
        
        Args:
            zip_path: Ruta al archivo ZIP
            
        Returns:
            Dict: Estadísticas del procesamiento
        """
        stats = {
            'zip_processed': False,
            'excel_extracted': False,
            'transferencias_found': 0,
            'transferencias_processed': 0,
            'transferencias_inserted': 0,
            'transferencias_duplicadas': 0,
            'errors': 0
        }
        
        excel_path = None
        temp_dir = None
        
        try:
            # Extraer ZIP
            excel_path = self.extract_zip(zip_path)
            if not excel_path:
                stats['errors'] += 1
                return stats
            
            stats['zip_processed'] = True
            stats['excel_extracted'] = True
            
            # Leer Excel
            transferencias = self.read_excel(excel_path, start_row=9)
            stats['transferencias_found'] = len(transferencias)
            
            if not transferencias:
                logger.warning("No se encontraron transferencias en el Excel")
                return stats
            
            # Procesar cada transferencia
            for transferencia in transferencias:
                try:
                    stats['transferencias_processed'] += 1

                    # Obtener idUnico para validación (campo interno _idUnico)
                    idUnico = transferencia.get('_idUnico')

                    # Validar que tenga idUnico
                    if not idUnico:
                        logger.warning(f"Transferencia sin idUnico, omitiendo: {transferencia.get('beneficiario')}")
                        stats['errors'] += 1
                        continue

                    # Verificar si es duplicado por idUnico antes de insertar
                    existing = self.supabase_client.get_movimiento_by_idunico(idUnico)
                    if existing:
                        stats['transferencias_duplicadas'] += 1
                        logger.warning(f"Transferencia duplicada (idUnico): {idUnico}")
                        continue

                    # Eliminar _idUnico antes de insertar (es solo para validación interna)
                    transferencia_sin_validacion = {k: v for k, v in transferencia.items() if k != '_idUnico'}

                    # Insertar en Supabase
                    if self.supabase_client.insert_movimiento_bancario(transferencia_sin_validacion):
                        stats['transferencias_inserted'] += 1
                        logger.info(f"Transferencia insertada: idUnico={idUnico}, Rastreo={transferencia.get('rastreo')}, Referencia={transferencia.get('referencia')}")
                    else:
                        stats['errors'] += 1
                        logger.error(f"Error al insertar transferencia: idUnico={idUnico}")
                        
                except Exception as e:
                    logger.error(f"Error al procesar transferencia: {str(e)}")
                    stats['errors'] += 1
                    continue
            
            return stats
            
        except Exception as e:
            logger.error(f"Error al procesar archivo ZIP: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            stats['errors'] += 1
            return stats
        
        finally:
            # Limpiar archivos temporales
            if excel_path:
                try:
                    temp_dir = os.path.dirname(excel_path)
                    if temp_dir and os.path.exists(temp_dir):
                        import shutil
                        shutil.rmtree(temp_dir)
                        logger.info("Archivos temporales eliminados")
                except Exception as e:
                    logger.warning(f"No se pudieron eliminar archivos temporales: {e}")


def main():
    """Función principal"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Procesar archivo ZIP con transferencias Excel')
    parser.add_argument('zip_file', help='Ruta al archivo ZIP a procesar')
    parser.add_argument('--start-row', type=int, default=9, 
                       help='Fila donde empiezan los datos (default: 9)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.zip_file):
        logger.error(f"El archivo no existe: {args.zip_file}")
        return 1
    
    logger.info("=" * 60)
    logger.info("PROCESADOR DE TRANSFERENCIAS DESDE EXCEL")
    logger.info("=" * 60)
    
    processor = ExcelTransferProcessor()
    stats = processor.process_zip_file(args.zip_file)
    
    logger.info("=" * 60)
    logger.info("RESUMEN DEL PROCESAMIENTO")
    logger.info("=" * 60)
    logger.info(f"ZIP procesado: {'SI' if stats['zip_processed'] else 'NO'}")
    logger.info(f"Excel extraido: {'SI' if stats['excel_extracted'] else 'NO'}")
    logger.info(f"Transferencias encontradas: {stats['transferencias_found']}")
    logger.info(f"Transferencias procesadas: {stats['transferencias_processed']}")
    logger.info(f"Transferencias insertadas: {stats['transferencias_inserted']}")
    logger.info(f"Transferencias duplicadas: {stats['transferencias_duplicadas']}")
    logger.info(f"Errores: {stats['errors']}")
    logger.info("=" * 60)
    
    if stats['errors'] > 0:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())

